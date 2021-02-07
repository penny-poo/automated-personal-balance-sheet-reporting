
import openpyxl as xl
import datetime


old_name = input('What is the name of the Excel workbook to be reformatted? (Include file format)\n')
new_name = input('What do you want the new workbook to be called? (Include file format)\nWARNING: This will overwrite any file with the existing name\n')

'Balance sheet untabbed.xlsx'

old_wb = xl.load_workbook(old_name)

def find_sums(sheet):
    """
    Tallies the total assets and total liabilities for each person.
    
    RETURNS:
      Tuple of assets and liabilities.
    """
    pos = 0
    neg = 0
    for row in sheet:
        if row[-1] > 0:
            pos += row[-1]
        else:
            neg += row[-1]
    return pos, neg

def get_rows(wb):
    """
    Gets all values from first worksheet of workbook.

    Parameters
    ----------
    wb : excel workbook

    Returns
    -------
    TUPLE
        Tuple of rows.

    """
    return list(wb.worksheets[0].values)

def format_names(rows):
    """
    Takes the workbook data and formats the names to be displayed correctly later.
    """
    frows = []
    for row in rows:
        new_row = list(row)
        new_row[1] = new_row[1].strip()
        new_row[2] = new_row[2].strip()
        if type(new_row[4]) == type(datetime.datetime(2000, 10, 20)):
            new_row[4] = new_row[4].date()
        frows.append(new_row)
    return frows

def get_names_from_rows(frows):
    """
    Gets names from workbook and returns a list

    Returns
    -------
    names : LIST
        ['first last', 'first last'] ... all are stripped and lowered.

    """
    names = []
    for row in frows:
        name = row[2] + ', ' + row[1]
        if name not in names:
            names.append(name)
    return sorted(names)



def create_sheets_from_names(names, frows):
    """
    Returns a new sheet for each client.
    """
    new_sheets = []
    for name in names:
        sheet = []
        last, first = name.split(', ')
        for row in frows:
            if first in row and last in row:
                sheet.append(row)
        sheet = sorted(sheet, key=lambda x: x[-1], reverse = True)
        new_sheets.append(sheet)
    return new_sheets

def construct_new_wb(names, sheets):
    """
    Creates a new workbook with a sheet for each separate client. Sorts and labels
    assets, liabilites, and net worth and presents it in an appealing way.
    """
    wb = xl.Workbook()
    
    # This format mimics the accounting format of excel.
    fmt_acct =  u'_($* #,##0.00_);[Red]_($* (#,##0.00);_($* -_0_0_);_(@'
    for i in range(len(names)):
        wb.create_sheet(names[i])
        wb[names[i]].append([names[i] + ' Balance Sheet/Net Worth Statement'])
        wb[names[i]].append(['Assets', 'First Name', 'Last Names', 'Account Type', 'As of Date', 'Account Balance'])
        
        # Formatting first two rows
        wb[names[i]]['A1'].font = xl.styles.Font(b=True)
        wb[names[i]]['A2'].font = xl.styles.Font(b=True, u='single')
        wb[names[i]]['B2'].font = xl.styles.Font(b=True, u='single')
        wb[names[i]]['C2'].font = xl.styles.Font(b=True, u='single')
        wb[names[i]]['D2'].font = xl.styles.Font(b=True, u='single')
        wb[names[i]]['E2'].font = xl.styles.Font(b=True, u='single')
        wb[names[i]]['F2'].font = xl.styles.Font(b=True, u='single')
        flag = False
        pos_sum, neg_sum = find_sums(sheets[i])
        
        # Formatting idiosyncratic labels.
        r = 3
        for j, row in enumerate(sheets[i]):
            if row[-1] < 0 and flag == False:
                wb[names[i]]['F' + str(r-1)].border = xl.styles.Border(bottom=xl.styles.borders.Side(color='00000000', border_style='medium'))
                wb[names[i]].append(['','','','','subtotal',pos_sum])
                wb[names[i]]['F' + str(r)].number_format = fmt_acct
                wb[names[i]]['E' + str(r)].font = xl.styles.Font(b=True)
                wb[names[i]]['F' + str(r)].font = xl.styles.Font(b=True)
                r += 1
                wb[names[i]].append(['Liabilities'])
                wb[names[i]]['A' + str(r)].font = xl.styles.Font(b=True)
                r += 1
                flag = True
            wb[names[i]].append(row)
            if flag == True:
                wb[names[i]]['F' + str(r)].font = xl.styles.Font(color ='00FF0000',)
            wb[names[i]]['F' + str(r)].number_format = fmt_acct
            r += 1
            if j == len(sheets[i])-1:
                if flag == False:
                    wb[names[i]]['F' + str(r-1)].border = xl.styles.Border(bottom=xl.styles.borders.Side(color='00000000', border_style='medium'))
                    wb[names[i]].append(['','','','','subtotal',pos_sum])
                    wb[names[i]]['F' + str(r)].number_format = fmt_acct
                    wb[names[i]]['F' + str(r)].font = xl.styles.Font(b=True)
                    wb[names[i]]['E' + str(r)].font = xl.styles.Font(b=True)
                    r += 1
                elif flag == True:
                    wb[names[i]]['F' + str(r-1)].border = xl.styles.Border(bottom=xl.styles.borders.Side(color='00000000', border_style='medium'))
                    wb[names[i]].append(['','','','','subtotal',neg_sum])
                    wb[names[i]]['F' + str(r)].number_format = fmt_acct
                    wb[names[i]]['F' + str(r)].font = xl.styles.Font(b=True, color='00FF0000')
                    wb[names[i]]['E' + str(r)].font = xl.styles.Font(b=True)
                    r += 1
                wb[names[i]].append([])
                r += 1
                wb[names[i]].append(['','','','','Net Worth', pos_sum + neg_sum])
                wb[names[i]]['F' + str(r)].number_format = fmt_acct
                wb[names[i]]['E' + str(r)].font = xl.styles.Font(b=True)
                wb[names[i]]['F' + str(r)].font = xl.styles.Font(b=True)
                # for number in range(3, r):
                #     wb[names[i]]['F' + str(number)].style = 'Currency'
    del wb['Sheet']
    return wb

# Runs script. All it needs is for the desired file to be in the same directory.
rows = get_rows(old_wb)
frows = format_names(rows)
names = get_names_from_rows(frows)
sheets = create_sheets_from_names(names, frows)
wb = construct_new_wb(names, sheets)
wb.save(filename = new_name)
